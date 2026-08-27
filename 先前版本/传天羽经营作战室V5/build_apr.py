#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从销售分析表2026提取 APR 产品结构（四大主机占比 + 配件占比）2026-8 真实数据"""
import json, os
from openpyxl import load_workbook

SRC = os.path.expanduser("~/Desktop/APR项目/运营数据分析表/26年/销售分析表2026.xlsx")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "apr_struct.json")

wb = load_workbook(SRC, data_only=True, read_only=True)

struct = {"month": "2026-8", "host": None, "acc": None, "stores_host": [], "stores_acc": []}

# ===== 四大主机占比：整体 + 各门店 =====
ws = wb["四大主机占比"]
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row or str(row[2] or "").strip() != "2026-8":
        continue
    store = str(row[1] or "").strip()
    d = {
        "store": store,
        "total_amount": round(row[3] or 0, 2),   # 线下销售总金额(万)
        "total_qty": row[4] or 0,
        "host_amount": round(row[5] or 0, 2),
        "host_qty": row[6] or 0,
        "iphone": {"amount": round(row[7] or 0, 2), "pct": round((row[8] or 0) * 100, 1), "qty": row[9] or 0},
        "ipad": {"amount": round(row[11] or 0, 2), "pct": round((row[12] or 0) * 100, 1), "qty": row[13] or 0},
        "watch": {"amount": round(row[15] or 0, 2), "pct": round((row[16] or 0) * 100, 1), "qty": row[17] or 0},
        "mac": {"amount": round(row[19] or 0, 2), "pct": round((row[20] or 0) * 100, 1), "qty": row[21] or 0},
    }
    if store == "整体":
        struct["host"] = d
    else:
        struct["stores_host"].append(d)

# ===== 配件占比分析：整体 + 各门店 =====
ws2 = wb["配件占比分析"]
for row in ws2.iter_rows(min_row=4, values_only=True):
    if not row or str(row[2] or "").strip() != "2026-8":
        continue
    store = str(row[1] or "").strip()
    d = {
        "store": store,
        "acc_amount": round(row[3] or 0, 2),   # 第三方配件销售金额(万)
        "acc_qty": row[5] or 0,
        "phone": {"amount": round(row[6] or 0, 2), "pct": round((row[7] or 0) * 100, 1), "qty": row[8] or 0},
        "pad": {"amount": round(row[10] or 0, 2), "pct": round((row[11] or 0) * 100, 1), "qty": row[12] or 0},
        "watch": {"amount": round(row[14] or 0, 2), "pct": round((row[15] or 0) * 100, 1), "qty": row[16] or 0},
        "other": {"amount": round(row[22] or 0, 2), "pct": round((row[23] or 0) * 100, 1), "qty": row[24] or 0},
    }
    if store == "整体":
        struct["acc"] = d
    else:
        struct["stores_acc"].append(d)

with open(OUT, "w") as f:
    json.dump(struct, f, ensure_ascii=False, indent=1)

if struct["host"]:
    h = struct["host"]
    print("四大主机(2026-8 线下): 总销售", h["total_amount"], "万, 主机", h["host_amount"], "万")
    print("  iPhone", h["iphone"]["amount"], f"万({h['iphone']['pct']}%) | iPad", h["ipad"]["amount"], f"万({h['ipad']['pct']}%) | Watch", h["watch"]["amount"], f"万({h['watch']['pct']}%) | Mac", h["mac"]["amount"], f"万({h['mac']['pct']}%)")
if struct["acc"]:
    a = struct["acc"]
    print("配件(2026-8 线下): 第三方配件", a["acc_amount"], "万, 手机配件", a["phone"]["amount"], f"万({a['phone']['pct']}%) | 平板", a["pad"]["amount"], f"万({a['pad']['pct']}%) | 手表", a["watch"]["amount"], f"万({a['watch']['pct']}%)")
print("门店主机数据:", len(struct["stores_host"]), "家, 门店配件数据:", len(struct["stores_acc"]), "家")
