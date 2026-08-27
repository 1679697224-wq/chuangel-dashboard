#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""解析桌面《库存分析表_260807.xlsx》→ inv_data.json
- 金额口径为主（单位统一为万元），件数为辅
- 过滤虚拟品（发票/优惠券/差价等非实物）
- 库龄预警：90天+ 高库龄标记
"""
import json, os
from openpyxl import load_workbook

SRC = os.path.expanduser("~/Desktop/库存分析表_260807.xlsx")
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "inv_data.json")

VIRTUAL_KW = ["发票", "优惠券", "差价", "运费", "包装", "服务", "延保", "安装", "回收", "以旧换新", "赠品", "卡券"]

def is_virtual(name, brand, cat, amt):
    if not name:
        return False
    n = str(name)
    if any(k in n for k in VIRTUAL_KW):
        return True
    # 0金额 + 非实物品牌/分类 = 虚拟品（储值卡/权益/店铺费用/SVIP等）
    b = str(brand or "")
    c = str(cat or "")
    if amt <= 0 and (b in ("", "其他") or c in ("其他", "平台优惠券", "商场优惠券", "虚拟")):
        return True
    return False

wb = load_workbook(SRC, data_only=True, read_only=True)
ws = wb["总览"]
rows = list(ws.iter_rows(values_only=True))

def cell(r, c):
    try:
        return rows[r][c]
    except Exception:
        return None

data = {
    "source": "库存分析表_260807.xlsx（商务每日分析）",
    "category": [], "warehouse": [],
    "aging": {}, "aging_risk": [],
    "virtual_removed": {"count": 0, "amount": 0, "items": []},
    "total": {"qty": 0, "amount": 0},
}

# 分类结构（1-indexed 6-10: 合计/苹果主机/原装配件/舒尔/第三方配件）
for r in range(5, 10):
    name = str(cell(r, 1) or "").strip()
    if not name or name == "合计":
        continue
    data["category"].append({
        "name": name,
        "qty": cell(r, 4) or 0,
        "amount": round(cell(r, 5) or 0, 2),
        "pct": round((cell(r, 6) or 0) * 100, 1),
    })

# 仓位分布（1-indexed 15-21）
for r in range(14, 21):
    name = str(cell(r, 1) or "").strip()
    if not name or name == "合计":
        continue
    data["warehouse"].append({
        "name": name,
        "qty": cell(r, 4) or 0,
        "amount": round(cell(r, 5) or 0, 2),
        "pct": round((cell(r, 6) or 0) * 100, 1),
    })

# ===== 机况拆分：各渠道 常规机/样机/残次机（总览sheet 库龄分析区，金额万） =====
machine = {}
last_ch = ""
for r in range(36, 51):
    ch_name = str(cell(r, 1) or "").strip()
    sub = str(cell(r, 3) or "").strip()
    if ch_name:
        last_ch = ch_name
    if last_ch and sub:
        machine.setdefault(last_ch, {})[sub] = round(cell(r, 4) or 0, 2)
data["machine"] = machine

# ===== APR最低库存要求 vs 当前库存（零库存/缺货覆盖预警） =====
ws_min = wb["APR最低库存要求"]
min_rows = {}
for row in ws_min.iter_rows(min_row=2, values_only=True):
    if not row or not row[3]:
        continue
    min_rows[str(row[3]).strip()] = {
        "desc": str(row[4] or "").strip()[:40],
        "min_qty": row[5] or 0,
        "unit_price": row[6] or 0,
        "min_amount": row[7] or 0,
    }

# ===== 库龄明细表：按库龄分段汇总 + 高库龄预警（金额单位=元→万） =====
ws2 = wb["库龄"]
cur_stock = {}
for row in ws2.iter_rows(min_row=3, values_only=True):
    if not row or not row[1]:
        continue
    gno = str(row[1]).strip()
    gname = str(row[2] or "").strip()
    brand, cat = str(row[6] or ""), str(row[7] or "")
    amt = (row[12] or 0) / 10000.0
    if is_virtual(gname, brand, cat, amt):
        continue
    if gno not in cur_stock:
        cur_stock[gno] = 0
    cur_stock[gno] += int(row[11] or 0)
aging_rows = []
buckets = {"0~30天": 0, "30~60天": 0, "60~90天": 0, "90~180天": 0, "180~360天": 0, "360天以上": 0}
for row in ws2.iter_rows(min_row=3, values_only=True):
    if not row or not row[1]:
        continue
    warehouse, gno, gname, spec, unit, barcode, brand, cat = row[0:8]
    age = row[10] or 0
    qty = row[11] or 0
    amt = (row[12] or 0) / 10000.0  # 元 -> 万
    if is_virtual(gname, brand, cat, amt):
        data["virtual_removed"]["count"] += qty
        data["virtual_removed"]["amount"] += amt
        if len(data["virtual_removed"]["items"]) < 8:
            data["virtual_removed"]["items"].append(str(gname))
        continue
    if not qty:
        continue
    aging_rows.append({
        "wh": str(warehouse).replace("[", "").replace("]", "")[:6],
        "name": str(gname)[:26], "brand": str(brand or ""),
        "cat": str(cat or ""), "age": int(age),
        "qty": int(qty), "amount": round(amt, 2),
    })
    if age <= 30: buckets["0~30天"] += amt
    elif age <= 60: buckets["30~60天"] += amt
    elif age <= 90: buckets["60~90天"] += amt
    elif age <= 180: buckets["90~180天"] += amt
    elif age <= 360: buckets["180~360天"] += amt
    else: buckets["360天以上"] += amt

data["aging"] = {k: round(v, 2) for k, v in buckets.items()}

data["total"]["qty"] = sum(x["qty"] for x in aging_rows)
data["total"]["amount"] = round(sum(x["amount"] for x in aging_rows), 2)

risk = [x for x in aging_rows if x["age"] >= 90]
risk.sort(key=lambda x: -x["amount"])
data["aging_risk"] = risk[:15]
data["aging_risk_count"] = len(risk)
data["aging_risk_amount"] = round(sum(x["amount"] for x in risk), 2)
data["aging_risk_pct"] = round(data["aging_risk_amount"] / data["total"]["amount"] * 100, 1) if data["total"]["amount"] else 0

severe = [x for x in aging_rows if x["age"] >= 360]
data["severe_count"] = len(severe)
data["severe_amount"] = round(sum(x["amount"] for x in severe), 2)

data["virtual_removed"]["amount"] = round(data["virtual_removed"]["amount"], 2)

# 最低库存覆盖预警
coverage = []
for mpn, req in min_rows.items():
    cur = cur_stock.get(mpn, 0)
    if cur < int(req["min_qty"] or 0):
        coverage.append({
            "mpn": mpn, "desc": req["desc"],
            "min_qty": int(req["min_qty"] or 0), "cur_qty": cur,
            "gap": int(req["min_qty"] or 0) - cur,
            "status": "零库存" if cur == 0 else "低于最低要求",
        })
coverage.sort(key=lambda x: -x["gap"])
data["coverage"] = coverage[:20]
data["coverage_count"] = len(coverage)

with open(OUT, "w") as f:
    json.dump(data, f, ensure_ascii=False, indent=1)

print("inv_data.json 生成完成")
print("总库存(剔除虚拟品):", data["total"]["qty"], "件 /", data["total"]["amount"], "万")
print("分类:", [(c["name"], c["amount"]) for c in data["category"]])
print("仓位:", [(w["name"], w["amount"]) for w in data["warehouse"]])
print("库龄分布(万):", data["aging"])
print("剔除虚拟品:", data["virtual_removed"]["count"], "件 /", data["virtual_removed"]["amount"], "万")
print("高库龄90天+: ", data["aging_risk_count"], "项 /", data["aging_risk_amount"], "万 占", data["aging_risk_pct"], "%")
print("严重360天+:", data["severe_count"], "项 /", data["severe_amount"], "万")
