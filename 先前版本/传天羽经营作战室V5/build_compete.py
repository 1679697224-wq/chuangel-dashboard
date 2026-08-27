#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""解析《京东平台竞争产品信息统计.xlsx》→ compete.json + compete_compare.json
compete_compare: 按型号分组对比（竞品店铺覆盖款数+链接 vs 我们）
"""
import json, os, re
from openpyxl import load_workbook

SRC = os.path.expanduser("~/Desktop/7月/京东平台竞争产品信息统计.xlsx")
BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, "compete.json")
OUT2 = os.path.join(BASE, "compete_compare.json")

wb = load_workbook(SRC, data_only=True)
ws = wb["Sheet1"]

shops = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row or not row[0]:
        continue
    shop = str(row[0]).strip()
    prod = str(row[1] or "").strip()
    link = str(row[2] or "").strip()
    shops.setdefault(shop, []).append({"product": prod, "link": link})

compete = {
    "source": "京东平台竞争产品信息统计.xlsx（7月）",
    "platform": "京东",
    "total_products": sum(len(v) for v in shops.values()),
    "shops": [{"name": k, "products": v} for k, v in shops.items()],
}
with open(OUT, "w") as f:
    json.dump(compete, f, ensure_ascii=False, indent=1)

# ===== 型号分组（对比用） =====
def model_of(name):
    n = str(name)
    if "iPhone 17 Pro Max" in n: return "iPhone 17 Pro Max"
    if "iPhone 17 Pro" in n: return "iPhone 17 Pro"
    if "iPhone 17" in n: return "iPhone 17"
    if "iPad" in n: return "iPad"
    if "Watch" in n or "watch" in n: return "Apple Watch"
    if "AirPods" in n: return "AirPods"
    if "Mac" in n: return "Mac"
    return "其他"

def cap_color(name):
    """提取容量（256G/512G等）和颜色"""
    cap = re.search(r"(\d+G[BT])", str(name))
    color = ""
    for kw in ["蓝色","黑色","白色","紫色","绿色","灰色","橙色","粉色","原色","深蓝色","鼠尾草绿","薰衣草紫","星宇橙","淡桃粉","沙漠钛","钛金属"]:
        if kw in str(name):
            color = kw; break
    return (cap.group(1) if cap else "—"), (color or "—")

compare = {
    "platform": "京东",
    "our_shops": ["羽通｜京东", "啟韬｜苏宁"],
    "models": [],
    "summary": {},
}
model_map = {}
for shop, prods in shops.items():
    for p in prods:
        m = model_of(p["product"])
        if m == "其他":
            continue
        model_map.setdefault(m, {})
        model_map[m].setdefault(shop, [])
        model_map[m][shop].append(p)

# 排序型号：iPhone 17 Pro Max > Pro > 17 > 其他
order = {"iPhone 17 Pro Max": 0, "iPhone 17 Pro": 1, "iPhone 17": 2, "iPad": 3, "Apple Watch": 4, "AirPods": 5, "Mac": 6}
for m in sorted(model_map.keys(), key=lambda x: order.get(x, 9)):
    shops_data = {}
    for shop, items in model_map[m].items():
        caps = {}
        for it in items:
            cap, color = cap_color(it["product"])
            caps.setdefault(cap, {"count": 0, "link": "", "colors": set()})
            caps[cap]["count"] += 1
            caps[cap]["link"] = it["link"]
            if color != "—":
                caps[cap]["colors"].add(color)
        shops_data[shop] = {
            "count": len(items),
            "caps": [{"cap": k, "count": v["count"], "link": v["link"], "colors": list(v["colors"])[:4]} for k, v in sorted(caps.items())],
        }
    compare["models"].append({"model": m, "shops": shops_data, "total": sum(v["count"] for v in shops_data.values())})

compare["summary"] = {
    "total": {s["name"]: len(s["products"]) for s in compete["shops"]},
    "models_count": len(compare["models"]),
}
with open(OUT2, "w") as f:
    json.dump(compare, f, ensure_ascii=False, indent=1)

print(f"竞品对比生成: {len(compete['shops'])} 店 / {compete['total_products']} 款")
for m in compare["models"]:
    detail = " | ".join(f"{s}:{v['count']}款" for s, v in m["shops"].items())
    print(f"  {m['model']}: {detail}")
